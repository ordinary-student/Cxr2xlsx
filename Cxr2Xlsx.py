# -*- coding:utf-8 -*-
import os
import random
import sys
import time as _time
from openpyxl import Workbook
from PySide2.QtGui import QColor, QDragEnterEvent, QDropEvent, QIcon, QKeySequence, QPixmap, QTextCursor, Qt
from PySide2.QtWidgets import QApplication, QCheckBox, QGroupBox, QHBoxLayout,  QMainWindow, QMessageBox, QPushButton, QRadioButton, QStyleFactory, QTextEdit, QVBoxLayout, QWidget


version = 'v1.0.0'
last_update = '2021-07-04'


class MVariable(object):
    '''自定义变量类'''

    def __init__(self, var_type, address, name, value):
        '''构造函数'''
        # 变量类型
        self.var_type = var_type
        # 变量地址
        self.address = address
        # 变量名称-注释
        self.name = name
        # 初始值
        self.value = value

    def toList(self):
        '''转为List'''
        return [self.var_type, self.address, self.name, self.value]

    @staticmethod
    def toMV(l: list):
        '''列表转为MVariable'''
        if len(l) == 5:
            return MVariable(l[0], l[1], l[2], l[4])
        else:
            return MVariable('', '', '', '')

    @staticmethod
    def get_type_name(address: str) -> str:
        '''根据地址获取变量类型'''
        # 字符开头
        if address[0].isalpha():
            return address[0]
        else:
            return 'IO'


class MTextEdit(QTextEdit):
    '''自定义TextEdit类'''

    def __init__(self, parent):
        '''构造函数'''
        super(MTextEdit, self).__init__(parent)
        # 允许拖放
        self.setAcceptDrops(True)

    def dragEnterEvent(self, e: QDragEnterEvent):
        '''拽入事件'''
        # 判断是否包含地址
        if e.mimeData().hasUrls():
            # 接受数据
            e.accept()
        else:
            # 忽略数据
            e.ignore()

    def dropEvent(self, e: QDropEvent):
        '''释放事件'''
        # 获取文字
        text = self.toPlainText()
        # 遍历所有文件地址
        for url in e.mimeData().urls():
            file_path = str(url.toLocalFile())
            # 判断文件类型
            if (os.path.isfile(file_path)) and (file_path.endswith('.cxr')) and (file_path not in text):
                self.append(file_path)
                # 移动光标到最底
                self.moveCursor(QTextCursor.End)


class Cxr2Xlsx(QMainWindow):
    '''Cxr2Xlsx类'''

    def __init__(self, app: QApplication):
        '''构造函数'''
        super(Cxr2Xlsx, self).__init__()
        self.app = app

        # 加载界面ui
        self.initUI()

    def initUI(self):
        '''加载界面ui'''
        # 窗体标题
        self.setWindowTitle('cxr文件转为xlsx文件')
        # 设置窗口图标
        self.setWindowIcon(self.generateIcon())
        # 窗体大小
        self.resize(700, 700)
        # 最小大小
        self.setMinimumSize(700, 400)
        # 窗口居中
        self.center()

        # 中央容器
        self.centerWidget = QWidget(self)
        # 参数设置区域
        self.groupBox_setting = QGroupBox(
            '设置转换模式并拖放cxr文件进来', self.centerWidget)

        # 输出模式单选按钮
        self.btn_mode1 = QRadioButton("单输出模式")
        self.btn_mode1.setToolTip("所有cxr文件只转换输出一个xlsx文件，但不同cxr文件转换成单独的sheet表")
        self.btn_mode1.setChecked(True)

        self.btn_mode2 = QRadioButton("多输出模式")
        self.btn_mode2.setToolTip("每个cxr文件都转换为单独的xlsx文件")

        # 水平布局
        self.horizontalLayout_btn = QHBoxLayout()
        self.horizontalLayout_btn.addWidget(self.btn_mode1)
        self.horizontalLayout_btn.addWidget(self.btn_mode2)
        self.horizontalLayout_btn.setStretch(0, 1)
        self.horizontalLayout_btn.setStretch(1, 1)

        # 窗口置顶按钮
        self.checkBox_top = QCheckBox(
            '窗口置顶', self.groupBox_setting)
        self.checkBox_top.stateChanged.connect(self.set_top)

        # 开始转换按钮
        self.pushButton_start = QPushButton(
            '开始转换', self.groupBox_setting)
        self.pushButton_start.setStyleSheet(
            "QPushButton:hover{color: white;background:green}")
        self.pushButton_start.clicked.connect(self.convert)

        # 清空按钮
        self.pushButton_clear = QPushButton(
            '清空', self.groupBox_setting)
        self.pushButton_clear.setStyleSheet(
            "QPushButton:hover{color: blue}")
        self.pushButton_clear.clicked.connect(self.clear)

        # 关于按钮
        self.pushButton_about = QPushButton(
            '关于', self.groupBox_setting)
        self.pushButton_about.setStyleSheet(
            "QPushButton:hover{color: green}")
        self.pushButton_about.setShortcut(QKeySequence('F1'))
        self.pushButton_about.clicked.connect(self.about)

        # 打开输出目录按钮
        self.pushButton_opendir = QPushButton(
            '打开输出目录', self.groupBox_setting)
        self.pushButton_opendir.setStyleSheet(
            "QPushButton:hover{color: green}")
        self.pushButton_opendir.setShortcut(QKeySequence('F2'))
        self.pushButton_opendir.clicked.connect(self.opendir)

        # 水平布局2
        self.horizontalLayout_btn2 = QHBoxLayout()
        self.horizontalLayout_btn2.addLayout(self.horizontalLayout_btn)
        self.horizontalLayout_btn2.addWidget(self.checkBox_top)
        self.horizontalLayout_btn2.addWidget(self.pushButton_start)
        self.horizontalLayout_btn2.addWidget(self.pushButton_clear)
        self.horizontalLayout_btn2.addWidget(self.pushButton_opendir)
        self.horizontalLayout_btn2.addWidget(self.pushButton_about)
        self.horizontalLayout_btn2.setStretch(0, 2)
        self.horizontalLayout_btn2.setStretch(1, 1)
        self.horizontalLayout_btn2.setStretch(2, 1)
        self.horizontalLayout_btn2.setStretch(3, 1)
        self.horizontalLayout_btn2.setStretch(4, 1)
        self.horizontalLayout_btn2.setStretch(5, 1)

        # 文件拖放区
        self.fileTextEdit = MTextEdit(self.groupBox_setting)
        self.fileTextEdit.setToolTip("将cxr文件拖放到此处")

        # 垂直布局
        self.verticalLayout_setting = QVBoxLayout(self.groupBox_setting)
        self.verticalLayout_setting.addLayout(self.horizontalLayout_btn2)
        self.verticalLayout_setting.addWidget(self.fileTextEdit)
        self.verticalLayout_setting.setStretch(0, 1)
        self.verticalLayout_setting.setStretch(1, 3)

        # 信息输出区域
        self.groupBox_output = QGroupBox('信息输出', self.centerWidget)

        # 信息输出区
        self.outputTextEdit = QTextEdit(self)
        self.outputTextEdit.setAcceptDrops(False)

        # 垂直布局
        self.verticalLayout_output = QVBoxLayout(self.groupBox_output)
        self.verticalLayout_output.addWidget(self.outputTextEdit)

        # 中央容器垂直布局
        self.verticalLayout_center = QVBoxLayout(self.centerWidget)
        self.verticalLayout_center.addWidget(self.groupBox_setting)
        self.verticalLayout_center.addWidget(self.groupBox_output)
        self.verticalLayout_center.setStretch(0, 2)
        self.verticalLayout_center.setStretch(1, 3)

        # 设置中央容器
        self.setCentralWidget(self.centerWidget)

    def generateIcon(self) -> QIcon:
        '''生成图标'''
        # 新建图标
        pixmap = QPixmap(256, 256)
        # 图标颜色
        r = random.randint(0, 255)
        g = random.randint(0, 255)
        b = random.randint(0, 255)
        # 填充颜色
        pixmap.fill(QColor(r, g, b))
        # 返回
        return QIcon(pixmap)

    def center(self):
        '''窗口居中显示'''
        screen = self.app.primaryScreen().geometry()
        size = self.geometry()
        self.move((screen.width() - size.width()) / 2,
                  (screen.height() - size.height()) / 2)

    def set_top(self):
        '''窗口置顶'''
        # 判断勾选状态
        if self.checkBox_top.isChecked():
            # 置顶
            self.setWindowFlags(Qt.WindowStaysOnTopHint)
        else:
            # 取消置顶
            self.setWindowFlags(Qt.Widget)
        # 显示窗口
        self.show()

    def about(self):
        '''关于'''
        # 显示弹窗
        QMessageBox.about(self, '关于',
                          "cxr文件转为xlsx文件<br>author：<a href='https://github.com/ordinary-student'>ordinary-student</a><br>版本：{}<br>Last-Update：{}<br>© Copyright {}".format(version, last_update, last_update[0:4]))

    def nowtime(self, mode: int = 0) -> str:
        '''当前时间'''
        if mode == 0:
            return str(_time.strftime(
                '%Y-%m-%d %H:%M:%S', _time.localtime(_time.time())))
        else:
            return str(_time.strftime(
                '%Y%m%d-%H%M%S', _time.localtime(_time.time())))

    def log(self, message: str):
        '''输出信息'''
        mess = "<font color='orange'>[</font><font color='blue'>"+self.nowtime() + \
            "</font><font color='orange'>]</font><font color='green'>"+message+"</font>"
        self.outputTextEdit.append(mess)
        # 移动光标到最底
        self.outputTextEdit.moveCursor(QTextCursor.End)

    def clear(self):
        '''清空'''
        self.fileTextEdit.setText('')
        self.outputTextEdit.setText('')

    def write_to_single_sheet(self, xlsx_filename: str, sheetname: str, mv_list: list[MVariable]):
        '''将变量列表写入单个sheet表'''
        self.log('写入xlsx文件中...')
        # 新建工作簿文件
        wb = Workbook()
        # 新建表格
        sheet = wb.create_sheet(sheetname, index=0)
        # 遍历
        for mv in mv_list:
            # 添加一行
            sheet.append(mv.toList())
        # 判断文件名
        if not xlsx_filename.endswith('.xlsx'):
            xlsx_filename = xlsx_filename+'.xlsx'
        # 保存文件
        wb.save(xlsx_filename)
        self.log('写入xlsx文件完成！')

    def write_to_multi_sheet(self, xlsx_filename: str, sheetname_list: list[str], multi_mv_list: list[list]):
        '''将变量列表写入多个sheet表'''
        self.log('写入xlsx文件中...')
        # 判断
        if len(sheetname_list) != len(multi_mv_list):
            self.log('sheet表名称列表长度与cxr文件数量不符合！')
            return
        # 新建工作簿文件
        wb = Workbook()
        # 遍历
        for i in range(0, len(sheetname_list)):
            # 新建表格
            sheet = wb.create_sheet(sheetname_list[i], index=i)
            # 遍历
            for mv in multi_mv_list[i]:
                # 添加一行
                sheet.append(mv.toList())
        # 判断文件名
        if not xlsx_filename.endswith('.xlsx'):
            xlsx_filename = xlsx_filename+'.xlsx'
        # 保存文件
        wb.save(xlsx_filename)
        self.log('写入xlsx文件完成！')

    def get_list_from_cxr(self, file_path: str) -> list[MVariable]:
        '''从cxr文件读取变量列表'''
        # 判断文件类型
        if os.path.isfile(file_path) and file_path.endswith('.cxr'):
            self.log('读取cxr文件中...')
            # 读取文件
            with open(file_path, 'r') as f:
                # 按行读取
                lines = f.readlines()
            # 获取索引
            i = 0
            index: list[int] = []
            for i in range(0, len(lines)):
                if 'SYMBOL' in lines[i]:
                    index.append(i)
            # 截取变量部分
            content = lines[index[0]+1:index[1]]
            line3: list[str] = []
            for k in content:
                # 去除空格
                m = k[5:-2].replace('\t', '@#@')
                line3.append(m)
            # 结果集
            mv_list: list[MVariable] = []
            # 分割形成列表
            for n in line3:
                g = n.split('@#@')
                # 创建自定义变量
                mv = MVariable.toMV(g)
                mv_list.append(mv)
            self.log('此cxr文件获取变量列表完成！')
            # 返回
            return mv_list

    def convert(self):
        '''开始转换'''
        # 获取文字
        text = self.fileTextEdit.toPlainText()
        # 判断
        if len(text) <= 0:
            self.log('没有cxr文件！')
            return
        # 开始转换
        self.log('**********开始转换**********')
        # 分割
        filepath_list = text.split('\n')
        #
        multi_mv_list: list[list] = []
        sheetname_list: list[str] = []
        filename_list: list[str] = []
        i = 1
        try:
            # 遍历所有文件
            for filepath in filepath_list:
                # 判断路径不为空
                if len(filepath) != 0:
                    self.log('----------正在处理第{}个cxr文件----------'.format(i))
                    # 获取变量集
                    mv_list = self.get_list_from_cxr(filepath)
                    multi_mv_list.append(mv_list)
                    # 获取表名
                    sheetname = MVariable.get_type_name(mv_list[0].address)
                    sheetname_list.append(sheetname)

                    #  多输出模式
                    if self.btn_mode2.isChecked():
                        # 文件名
                        filename = '变量表-{}.xlsx'.format(sheetname)
                        filename_list.append(os.path.abspath(filename))
                        # 写入xlsx文件
                        self.write_to_single_sheet(
                            filename, sheetname, mv_list)

                    i = i+1
            #  单输出模式
            if self.btn_mode1.isChecked():
                # 文件名
                filename = '变量表-{}.xlsx'.format(self.nowtime(1))
                filename_list.append(os.path.abspath(filename))
                # 写入xlsx文件
                self.write_to_multi_sheet(
                    filename, sheetname_list, multi_mv_list)
            # 转换完成
            self.log('转换完成！输出文件位于：{}'.format('\n'.join(filename_list)))
            self.log('**********转换完成！**********')
            self.outputTextEdit.append('\n')
        except:
            self.log('转换出错！')
            self.outputTextEdit.append('\n')
            return

    def opendir(self):
        '''打开输出目录'''
        os.startfile(os.getcwd())


if __name__ == '__main__':
    # 创建应用
    app = QApplication(sys.argv)
    # 设置界面风格
    app.setStyle(QStyleFactory.create('Fusion'))
    # 创建窗口
    mainWindow = Cxr2Xlsx(app)
    # 显示
    mainWindow.show()
    # 退出
    sys.exit(app.exec_())
